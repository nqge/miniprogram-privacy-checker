#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 自动填写工具（简化版）
从检查结果文本文件中提取数据，自动填写到 Excel 模板中
"""

import re
from pathlib import Path
from typing import Dict, List
import sys

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[-] 错误: 未安装 openpyxl 库")
    print("[*] 请运行: pip install openpyxl")
    sys.exit(1)


class SimpleExcelFiller:
    """简化的 Excel 填写工具"""

    def __init__(self, base_path: str = None):
        """初始化"""
        if base_path:
            self.base_path = Path(base_path)
        else:
            self.base_path = Path(__file__).parent.parent

        # 查找 Excel 文件（按大小）
        xlsx_files = list(self.base_path.glob("*.xlsx"))
        xlsx_files.sort(key=lambda f: f.stat().st_size)

        if len(xlsx_files) >= 2:
            # 较小的是自评估表，较大的是权限确认单
            self.assessment_excel = xlsx_files[0]
            self.confirmation_excel = xlsx_files[1]
            print(f"[+] 找到自评估表 ({self.assessment_excel.stat().st_size} bytes)")
            print(f"[+] 找到权限确认单 ({self.confirmation_excel.stat().st_size} bytes)")
        else:
            print("[-] 错误: 未找到足够的 Excel 文件")
            sys.exit(1)

    def parse_permission_txt(self, txt_file: str) -> List[Dict]:
        """解析权限确认单文本文件"""
        permissions = []

        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # 查找表格开始
        for i, line in enumerate(lines):
            if '序号' in line and '权限组' in line and '敏感权限名称' in line:
                table_start = i + 2  # 跳过分隔线
                break
        else:
            print("[-] 未找到权限表格")
            return []

        # 解析每一行
        for line in lines[table_start:]:
            line = line.strip()
            if not line or line.startswith('-') or line.startswith('='):
                continue

            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 6:
                try:
                    applied = parts[3].replace('✅', '').replace('❌', '').strip()
                    permissions.append({
                        'name': parts[2],
                        'applied': applied,
                        'function': parts[4],
                        'necessity': parts[5] if len(parts) > 5 else ''
                    })
                except:
                    continue

        print(f"[+] 解析到 {len(permissions)} 条权限记录")
        return permissions

    def fill_confirmation(self, txt_file: str):
        """填写权限确认单"""
        permissions = self.parse_permission_txt(txt_file)
        if not permissions:
            return

        wb = load_workbook(self.confirmation_excel)
        ws = wb.active

        # 表头在第 3 行
        header_row = 3
        filled = 0

        for perm in permissions:
            # 查找对应的行
            for row_idx in range(header_row + 1, ws.max_row + 1):
                name_cell = ws.cell(row_idx, 3).value
                if name_cell and perm['name'] in str(name_cell):
                    # 填写数据到第 4, 5, 6 列
                    ws.cell(row_idx, 4, perm['applied'])
                    ws.cell(row_idx, 5, perm['function'])
                    ws.cell(row_idx, 6, perm['necessity'])
                    filled += 1
                    break

        print(f"[+] 成功填写 {filled} 条记录")

        # 保存文件
        output_path = str(self.confirmation_excel)
        wb.save(output_path)
        print(f"[+] 文件已保存 (大小: {self.confirmation_excel.stat().st_size} bytes)")

    def parse_assessment_txt(self, txt_file: str) -> Dict:
        """解析自评估表文本文件"""
        with open(txt_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 提取评分
        score_match = re.search(r'合规评分[::：]\s*(\d+)/100', content)
        score = int(score_match.group(1)) if score_match else 0

        # 确定评估结果
        if score >= 90:
            result = '符合'
            description = '小程序隐私合规性优秀，符合相关法律法规要求'
        elif score >= 70:
            result = '基本符合'
            description = '小程序隐私合规性良好，有少量需要改进的地方'
        elif score >= 50:
            result = '需要改进'
            description = '小程序存在一些隐私合规问题，需要改进'
        else:
            result = '不符合'
            description = '小程序存在严重的隐私合规问题，需要立即整改'

        print(f"[+] 解析评估结果: {result}, 评分: {score}")

        return {
            'result': result,
            'description': description,
            'score': score
        }

    def fill_assessment(self, txt_file: str):
        """填写自评估表"""
        assessment = self.parse_assessment_txt(txt_file)

        wb = load_workbook(self.assessment_excel)
        ws = wb.active

        # 表头在第 1 行
        header_row = 1

        # 查找列索引
        result_col = 3
        desc_col = 4

        # 填写数据（从第 2 行开始）
        data_row = 2
        ws.cell(data_row, result_col, assessment['result'])
        ws.cell(data_row, desc_col, assessment['description'])

        print(f"[+] 成功填写评估结果")

        # 保存文件
        output_path = str(self.assessment_excel)
        wb.save(output_path)
        print(f"[+] 文件已保存 (大小: {self.assessment_excel.stat().st_size} bytes)")


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='自动填写 Excel 权限确认单和自评估表')
    parser.add_argument('-b', '--base-path', help='基础路径',
                       default='/root/.openclaw/workspace/skills/miniprogram-privacy')
    parser.add_argument('-r', '--result-dir', help='检查结果目录',
                       default='privacy_check_results')

    args = parser.parse_args()

    filler = SimpleExcelFiller(args.base_path)
    result_path = Path(args.result_dir)

    # 查找并填写权限确认单
    confirmation_txt = list(result_path.glob("*权限确认单*.txt"))
    if confirmation_txt:
        print(f"\n[*] 处理权限确认单: {confirmation_txt[0].name}")
        filler.fill_confirmation(str(confirmation_txt[0]))
    else:
        print("\n[!] 未找到权限确认单文本文件")

    # 查找并填写自评估表
    assessment_txt = list(result_path.glob("*自评估*.txt"))
    if assessment_txt:
        print(f"\n[*] 处理自评估表: {assessment_txt[0].name}")
        filler.fill_assessment(str(assessment_txt[0]))
    else:
        print("\n[!] 未找到自评估表文本文件")


if __name__ == '__main__':
    main()
