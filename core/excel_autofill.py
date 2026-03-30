#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动填写 Excel 权限确认单和自评估表
从检查结果文本文件中提取数据，自动填写到 Excel 模板中
"""

import re
import json
from pathlib import Path
from typing import Dict, List, Tuple
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[-] 警告: 未安装 openpyxl 库，将尝试使用 xlrd/xlwt")


class ExcelAutoFiller:
    """Excel 自动填写工具"""

    def __init__(self, base_path: str = None):
        """
        初始化 Excel 自动填写工具

        Args:
            base_path: 小程序检查工具的基础路径
        """
        if base_path:
            self.base_path = Path(base_path)
        else:
            self.base_path = Path(__file__).parent.parent

        # 查找 Excel 文件
        self.confirmation_excel = self._find_excel_file("权限确认单")
        self.assessment_excel = self._find_excel_file("自评估表")

        if self.confirmation_excel:
            print(f"[+] 找到权限确认单 (大小: {self.confirmation_excel.stat().st_size} bytes)")
        else:
            print("[-] 未找到权限确认单 Excel 文件")

        if self.assessment_excel:
            print(f"[+] 找到自评估表 (大小: {self.assessment_excel.stat().st_size} bytes)")
        else:
            print("[-] 未找到自评估表 Excel 文件")

    def _find_excel_file(self, keyword: str) -> Path:
        """查找包含关键字的 Excel 文件"""
        for xlsx_file in self.base_path.glob("*.xlsx"):
            # 尝试读取文件名（处理编码问题）
            try:
                if keyword in xlsx_file.name.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore'):
                    return xlsx_file
            except:
                # 如果编码失败，尝试直接匹配
                if keyword.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore') in str(xlsx_file):
                    return xlsx_file

        # 如果找不到，尝试按大小判断
        xlsx_files = list(self.base_path.glob("*.xlsx"))
        if len(xlsx_files) == 2:
            # 两个文件，较小的可能是权限确认单
            xlsx_files.sort(key=lambda f: f.stat().st_size)
            if "确认" in keyword or "confirmation" in keyword.lower():
                return xlsx_files[0]
            else:
                return xlsx_files[1]

        return None

    def parse_permission_confirmation(self, txt_file: str) -> List[Dict]:
        """
        解析权限确认单文本文件

        Args:
            txt_file: 权限确认单文本文件路径

        Returns:
            权限列表
        """
        permissions = []

        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # 查找表格开始位置
        table_start = -1
        for i, line in enumerate(lines):
            if '序号' in line and '权限组' in line and '敏感权限名称' in line:
                table_start = i + 2  # 跳过分隔线
                break

        if table_start == -1:
            print("[-] 未找到权限表格")
            return permissions

        # 解析表格数据
        for i in range(table_start, len(lines)):
            line = lines[i].strip()

            # 跳过空行和分隔线
            if not line or line.startswith('-') or line.startswith('='):
                continue

            # 解析每一行
            # 格式: 序号 | 权限组 | 敏感权限名称 | 小程序是否申请 | 对应业务功能 | 必要性说明
            parts = [p.strip() for p in line.split('|')]

            if len(parts) >= 6:
                try:
                    # 提取申请状态（是/否）
                    applied = parts[3].replace('✅', '').replace('❌', '').strip()
                    is_applied = '是' in applied

                    permission = {
                        'id': int(parts[0]),
                        'group': parts[1],
                        'name': parts[2],
                        'is_applied': is_applied,
                        'applied_text': applied,
                        'business_function': parts[4],
                        'necessity': parts[5] if len(parts) > 5 else ''
                    }
                    permissions.append(permission)
                except ValueError as e:
                    # 跳过无法解析的行
                    continue

        print(f"[+] 解析到 {len(permissions)} 条权限记录")
        return permissions

    def fill_confirmation_excel(self, txt_file: str, output_file: str = None) -> bool:
        """
        填写权限确认单 Excel

        Args:
            txt_file: 权限确认单文本文件路径
            output_file: 输出文件路径（可选，默认覆盖原文件）

        Returns:
            是否成功
        """
        if not self.confirmation_excel:
            print("[-] 错误: 未找到权限确认单 Excel 文件")
            return False

        if not EXCEL_SUPPORT:
            print("[-] 错误: 未安装 openpyxl 库")
            print("[*] 请运行: pip install openpyxl")
            return False

        # 解析文本文件
        permissions = self.parse_permission_confirmation(txt_file)
        if not permissions:
            print("[-] 错误: 未能解析权限数据")
            return False

        try:
            # 加载 Excel 文件
            wb = load_workbook(self.confirmation_excel)
            ws = wb.active

            # 查找表头位置
            header_row = -1
            col_indices = {}

            for row_idx in range(1, ws.max_row + 1):
                # 检查每一行是否包含"序号"
                has_header = False
                for col_idx in range(1, min(7, ws.max_column + 1)):  # 只检查前 6 列
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value and isinstance(cell.value, str) and '序号' in cell.value:
                        has_header = True
                        header_row = row_idx
                        print(f"[DEBUG] Found header at row {row_idx}, col {col_idx}")
                        break

                if has_header:
                    # 找到各列的索引
                    for col_idx in range(1, min(7, ws.max_column + 1)):
                        cell = ws.cell(row_idx, col_idx)
                        if cell.value and isinstance(cell.value, str):
                            cell_str = str(cell.value)
                            if '序号' in cell_str:
                                col_indices['id'] = col_idx
                            elif '权限组' in cell_str:
                                col_indices['group'] = col_idx
                            elif '敏感权限名称' in cell_str:
                                col_indices['name'] = col_idx

                    # 如果找不到列名，假设固定顺序
                    if 'id' not in col_indices:
                        col_indices['id'] = 1
                    if 'group' not in col_indices:
                        col_indices['group'] = 2
                    if 'name' not in col_indices:
                        col_indices['name'] = 3

                    # 后 3 列固定为 4, 5, 6
                    col_indices['applied'] = 4
                    col_indices['function'] = 5
                    col_indices['necessity'] = 6
                    break

            if header_row == -1:
                print("[-] 未找到表头行")
                return False

            print(f"[+] 找到表头在第 {header_row} 行")
            print(f"[+] 列索引: {col_indices}")

            # 填写数据
            filled_count = 0
            for perm in permissions:
                # 查找对应的行（通过权限名称匹配）
                row_found = False
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    # 检查权限名称列
                    if 'name' in col_indices:
                        name_cell = ws.cell(row_idx, col_indices['name']).value
                        if name_cell and perm['name'] in str(name_cell):
                            # 填写数据
                            if 'applied' in col_indices:
                                ws.cell(row_idx, col_indices['applied'], perm['applied_text'])
                            if 'function' in col_indices:
                                ws.cell(row_idx, col_indices['function'], perm['business_function'])
                            if 'necessity' in col_indices:
                                ws.cell(row_idx, col_indices['necessity'], perm['necessity'])

                            filled_count += 1
                            row_found = True
                            break

                if not row_found:
                    print(f"[!] 未找到权限: {perm['name']}")

            print(f"[+] 成功填写 {filled_count} 条记录")

            # 保存文件
            if output_file:
                save_path = output_file
            else:
                save_path = str(self.confirmation_excel)

            wb.save(save_path)
            print(f"[+] 文件已保存: {save_path}")

            return True

        except Exception as e:
            print(f"[-] 填写 Excel 失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def parse_self_assessment(self, txt_file: str) -> Dict:
        """
        解析自评估表文本文件

        Args:
            txt_file: 自评估表文本文件路径

        Returns:
            评估结果字典
        """
        assessments = {}

        with open(txt_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 提取评分
        score_match = re.search(r'合规评分[::：]\s*(\d+)/100', content)
        if score_match:
            assessments['score'] = int(score_match.group(1))

        # 提取总体评估
        if '优秀' in content or score_match and int(score_match.group(1)) >= 90:
            assessments['result'] = '符合'
            assessments['description'] = '小程序隐私合规性优秀，符合相关法律法规要求'
        elif '良好' in content or score_match and int(score_match.group(1)) >= 70:
            assessments['result'] = '基本符合'
            assessments['description'] = '小程序隐私合规性良好，有少量需要改进的地方'
        elif '一般' in content or score_match and int(score_match.group(1)) >= 50:
            assessments['result'] = '需要改进'
            assessments['description'] = '小程序存在一些隐私合规问题，需要改进'
        else:
            assessments['result'] = '不符合'
            assessments['description'] = '小程序存在严重的隐私合规问题，需要立即整改'

        # 提取问题列表
        issues = []
        issue_pattern = re.compile(r'[🔴⚠️❌]\s*(.+?)[::：]\s*(.+)', re.MULTILINE)
        for match in issue_pattern.finditer(content):
            issue_type = match.group(1).strip()
            issue_desc = match.group(2).strip()
            issues.append(f"{issue_type}: {issue_desc}")

        if issues:
            assessments['issues'] = issues

        print(f"[+] 解析评估结果: {assessments.get('result', '未知')}, 评分: {assessments.get('score', 'N/A')}")
        return assessments

    def fill_assessment_excel(self, txt_file: str, output_file: str = None) -> bool:
        """
        填写自评估表 Excel

        Args:
            txt_file: 自评估表文本文件路径
            output_file: 输出文件路径（可选，默认覆盖原文件）

        Returns:
            是否成功
        """
        if not self.assessment_excel:
            print("[-] 错误: 未找到自评估表 Excel 文件")
            return False

        if not EXCEL_SUPPORT:
            print("[-] 错误: 未安装 openpyxl 库")
            print("[*] 请运行: pip install openpyxl")
            return False

        # 解析文本文件
        assessments = self.parse_self_assessment(txt_file)
        if not assessments:
            print("[-] 错误: 未能解析评估数据")
            return False

        try:
            # 加载 Excel 文件
            wb = load_workbook(self.assessment_excel)
            ws = wb.active

            # 查找表头位置
            header_row = -1
            col_indices = {}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if any(cell and '评估结果' in str(cell) for cell in row if cell):
                    header_row = row_idx
                    # 找到各列的索引
                    for col_idx, cell in enumerate(row, 1):
                        if cell:
                            cell_str = str(cell)
                            if '评估结果' in cell_str:
                                col_indices['result'] = col_idx
                            elif '评估说明' in cell_str:
                                col_indices['description'] = col_idx
                    break

            if header_row == -1:
                print("[-] 未找到表头行")
                return False

            print(f"[+] 找到表头在第 {header_row} 行")
            print(f"[+] 列索引: {col_indices}")

            # 填写数据
            # 假设评估结果在第一行数据
            data_row = header_row + 1

            if 'result' in col_indices:
                ws.cell(data_row, col_indices['result'], assessments.get('result', ''))

            # 构建评估说明
            description_parts = [assessments.get('description', '')]
            if 'issues' in assessments:
                description_parts.append("\n主要问题:")
                description_parts.extend(assessments['issues'][:5])  # 最多显示5个问题

            if 'description' in col_indices:
                ws.cell(data_row, col_indices['description'], '\n'.join(description_parts))

            print(f"[+] 成功填写评估结果")

            # 保存文件
            if output_file:
                save_path = output_file
            else:
                save_path = str(self.assessment_excel)

            wb.save(save_path)
            print(f"[+] 文件已保存: {save_path}")

            return True

        except Exception as e:
            print(f"[-] 填写 Excel 失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def fill_all(self, result_dir: str, output_dir: str = None):
        """
        自动填写所有 Excel 文件

        Args:
            result_dir: 检查结果目录（包含 .txt 文件）
            output_dir: 输出目录（可选）
        """
        result_path = Path(result_dir)

        # 查找权限确认单文本文件
        confirmation_txt = list(result_path.glob("*权限确认单*.txt"))
        if confirmation_txt:
            print(f"\n[*] 处理权限确认单: {confirmation_txt[0].name}")
            self.fill_confirmation_excel(str(confirmation_txt[0]))
        else:
            print("\n[!] 未找到权限确认单文本文件")

        # 查找自评估表文本文件
        assessment_txt = list(result_path.glob("*自评估*.txt"))
        if assessment_txt:
            print(f"\n[*] 处理自评估表: {assessment_txt[0].name}")
            self.fill_assessment_excel(str(assessment_txt[0]))
        else:
            print("\n[!] 未找到自评估表文本文件")


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='自动填写 Excel 权限确认单和自评估表')
    parser.add_argument('-b', '--base-path', help='小程序检查工具基础路径',
                       default='/root/.openclaw/workspace/skills/miniprogram-privacy')
    parser.add_argument('-r', '--result-dir', help='检查结果目录（包含 .txt 文件）',
                       default='privacy_check_results')
    parser.add_argument('-o', '--output-dir', help='输出目录（可选）')
    parser.add_argument('--confirmation-txt', help='权限确认单文本文件路径')
    parser.add_argument('--assessment-txt', help='自评估表文本文件路径')

    args = parser.parse_args()

    # 创建自动填写工具
    filler = ExcelAutoFiller(args.base_path)

    # 如果指定了具体的文本文件
    if args.confirmation_txt:
        print("[*] 填写权限确认单...")
        filler.fill_confirmation_excel(args.confirmation_txt)

    if args.assessment_txt:
        print("[*] 填写自评估表...")
        filler.fill_assessment_excel(args.assessment_txt)

    # 否则自动查找并填写
    if not args.confirmation_txt and not args.assessment_txt:
        filler.fill_all(args.result_dir, args.output_dir)


if __name__ == '__main__':
    main()
