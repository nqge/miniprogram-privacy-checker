#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强版 AI Excel 自动填写工具
使用 AI 智能分析检查结果，自动更新权限确认单和自评估表
"""

import json
import re
from pathlib import Path
from typing import Dict, List, Any
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from .ai_agent_engine import AIAgentEngine

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EnhancedAIExcelFiller:
    """增强版 AI Excel 填充工具"""

    def __init__(self, base_path: str = None, enable_ai: bool = False):
        """
        初始化增强版 Excel 填充工具

        Args:
            base_path: 基础路径
            enable_ai: 是否启用 AI 分析
        """
        if not EXCEL_SUPPORT:
            raise ImportError("未安装 openpyxl 库，请运行: pip install openpyxl")

        self.enable_ai = enable_ai
        self.ai_engine = AIAgentEngine(enable_ai=enable_ai) if enable_ai else None

        if base_path:
            self.base_path = Path(base_path)
        else:
            self.base_path = Path(__file__).parent.parent

        # 查找 Excel 文件
        xlsx_files = list(self.base_path.glob("*.xlsx"))
        xlsx_files.sort(key=lambda f: f.stat().st_size)

        # 识别文件
        self.assessment_excel = None
        self.confirmation_excel = None

        for xlsx_file in xlsx_files:
            try:
                wb = load_workbook(xlsx_file)
                ws = wb.active

                # 检查第一行来判断文件类型
                first_cell = ws.cell(1, 1).value
                if first_cell and '类别' in str(first_cell):
                    # 这是自评估表
                    if not self.assessment_excel or xlsx_file.stat().st_size > self.assessment_excel.stat().st_size:
                        self.assessment_excel = xlsx_file
                elif first_cell and '小程序名称' in str(first_cell):
                    # 这是权限确认单
                    self.confirmation_excel = xlsx_file

                wb.close()
            except Exception as e:
                logger.warning(f"无法读取 {xlsx_file.name}: {e}")

        if self.assessment_excel:
            logger.info(f"[+] 找到自评估表 ({self.assessment_excel.stat().st_size} bytes)")
        if self.confirmation_excel:
            logger.info(f"[+] 找到权限确认单 ({self.confirmation_excel.stat().st_size} bytes)")

    def analyze_permission_for_excel(self, result_dir: str) -> Dict[str, Any]:
        """
        分析权限检查结果，生成权限确认单的业务功能和必要性说明

        Args:
            result_dir: 检查结果目录

        Returns:
            权限分析字典
        """
        logger.info("分析权限检查结果...")

        result_path = Path(result_dir)
        analysis = {
            'permissions': {},
            'overall_score': 0
        }

        # 读取权限检查结果
        perm_file = result_path / 'permission_check.json'
        if not perm_file.exists():
            logger.error("未找到权限检查结果")
            return analysis

        with open(perm_file, 'r', encoding='utf-8') as f:
            perm_data = json.load(f)

        # 读取 API 扫描结果
        api_file = result_path / 'api_scan.json'
        api_calls = {}
        if api_file.exists():
            with open(api_file, 'r', encoding='utf-8') as f:
                api_data = json.load(f)
                api_calls = api_data.get('api_calls', {})

        # 分析每个权限
        used_permissions = perm_data.get('used_permissions', {})

        for perm_name, perm_info in used_permissions.items():
            # 查找相关的 API 调用
            related_apis = [api for api in api_calls.keys() if perm_name.lower() in api.lower()]

            # 生成业务功能描述
            business_function = self._generate_business_function(perm_name, related_apis)

            # 生成必要性说明
            necessity = self._generate_necessity(perm_name, perm_info, related_apis)

            analysis['permissions'][perm_name] = {
                'business_function': business_function,
                'necessity': necessity,
                'usage_count': perm_info.get('count', 0),
                'related_apis': related_apis
            }

        # 计算总体评分
        total_perms = len(used_permissions)
        if total_perms > 0:
            declared_perms = len(perm_data.get('declared_permissions', {}))
            analysis['overall_score'] = int((declared_perms / total_perms) * 100)

        return analysis

    def _generate_business_function(self, perm_name: str, related_apis: List[str]) -> str:
        """生成业务功能描述"""
        business_functions = {
            'scope.userLocation': '用于获取用户地理位置信息，提供基于位置的个性化服务，如附近门店查找、配送服务等',
            'scope.camera': '用于拍摄照片或视频，支持用户上传图片、扫描二维码、拍摄证件等功能',
            'scope.writePhotosAlbum': '用于保存图片到用户相册，支持用户保存生成的报告、分享的图片等内容',
            'scope.record': '用于录制音频，支持语音输入、语音消息等功能',
            'scope.userInfo': '用于获取用户基本信息（昵称、头像等），提供个性化体验和用户识别'
        }

        # 如果有预定义的业务功能，直接返回
        if perm_name in business_functions:
            return business_functions[perm_name]

        # 否则根据 API 调用生成
        if 'getLocation' in str(related_apis):
            return '用于获取用户地理位置信息，提供基于位置的个性化服务'
        elif 'chooseImage' in str(related_apis) or 'chooseMedia' in str(related_apis):
            return '用于选择用户相册中的图片或视频，支持图片上传、编辑等功能'
        elif 'getUserInfo' in str(related_apis) or 'getUserProfile' in str(related_apis):
            return '用于获取用户基本信息，提供个性化服务和用户识别'
        else:
            return '用于支持核心业务功能，提升用户体验'

    def _generate_necessity(self, perm_name: str, perm_info: Dict, related_apis: List[str]) -> str:
        """生成必要性说明"""
        usage_count = perm_info.get('count', 0)

        if usage_count > 10:
            return f'必要 - 该权限为核心功能使用，在 {usage_count} 处调用，是实现业务目标的必要条件'
        elif usage_count > 5:
            return f'必要 - 该权限为主要功能使用，在 {usage_count} 处调用，对用户体验有重要影响'
        elif usage_count > 0:
            return f'较必要 - 该权限在 {usage_count} 处调用，用于提供增强功能'
        else:
            return '待评估 - 未检测到明显使用场景，建议进一步确认'

    def analyze_assessment_for_excel(self, result_dir: str) -> Dict[str, Any]:
        """
        分析自评估检查结果，生成评估结果和评估说明

        Args:
            result_dir: 检查结果目录

        Returns:
            自评估分析字典
        """
        logger.info("分析自评估检查结果...")

        result_path = Path(result_dir)
        analysis = {
            'assessments': {},
            'overall_score': 0
        }

        # 读取自评估结果
        assess_file = result_path / 'self_assessment.json'
        if not assess_file.exists():
            logger.error("未找到自评估结果")
            return analysis

        with open(assess_file, 'r', encoding='utf-8') as f:
            assess_data = json.load(f)

        assessments = assess_data.get('assessments', [])

        for item in assessments:
            item_id = item.get('id', '')
            category = item.get('category', '')
            description = item.get('description', '')
            result = item.get('result', '待评估')

            # 生成评估说明
            explanation = self._generate_assessment_explanation(item, result)

            analysis['assessments'][item_id] = {
                'category': category,
                'description': description,
                'result': result,
                'explanation': explanation
            }

        # 计算总体评分
        total = len(assessments)
        if total > 0:
            passed = sum(1 for a in assessments if a.get('result') == '符合')
            analysis['overall_score'] = int((passed / total) * 100)

        return analysis

    def _generate_assessment_explanation(self, item: Dict, result: str) -> str:
        """生成评估说明"""
        if result == '符合':
            return f'经检查，{item.get("description", "")}符合规范要求，已建立相应的保护机制和告知流程'
        elif result == '不符合':
            return f'经检查，{item.get("description", "")}不符合规范要求，存在合规风险，建议立即整改'
        elif result == '部分符合':
            return f'经检查，{item.get("description", "")}部分符合规范要求，建议进一步完善相关措施'
        else:
            return '待进一步评估'

    def fill_permission_confirmation(self, result_dir: str, output_path: str = None):
        """
        填写权限确认单 Excel

        Args:
            result_dir: 检查结果目录
            output_path: 输出文件路径
        """
        logger.info("填写权限确认单...")

        if not self.confirmation_excel:
            logger.error("未找到权限确认单模板")
            return

        # 分析权限检查结果
        analysis = self.analyze_permission_for_excel(result_dir)

        # 加载 Excel 文件
        wb = load_workbook(self.confirmation_excel)
        ws = wb.active

        # 查找数据起始行（通常是第3行）
        data_start_row = 3
        current_row = data_start_row

        # 遍历所有权限
        for perm_name, perm_data in analysis['permissions'].items():
            # 查找权限名称列（假设在第1列）
            if current_row < ws.max_row:
                # 更新业务功能（假设在第5列）
                ws.cell(current_row, 5, perm_data['business_function'])

                # 更新必要性说明（假设在第6列）
                ws.cell(current_row, 6, perm_data['necessity'])

                current_row += 1

        # 保存文件
        if not output_path:
            output_path = self.confirmation_excel

        wb.save(output_path)
        logger.info(f"权限确认单已更新: {output_path}")

    def fill_self_assessment(self, result_dir: str, output_path: str = None):
        """
        填写自评估表 Excel

        Args:
            result_dir: 检查结果目录
            output_path: 输出文件路径
        """
        logger.info("填写自评估表...")

        if not self.assessment_excel:
            logger.error("未找到自评估表模板")
            return

        # 分析自评估结果
        analysis = self.analyze_assessment_for_excel(result_dir)

        # 加载 Excel 文件
        wb = load_workbook(self.assessment_excel)
        ws = wb.active

        # 查找数据起始行（通常是第3行）
        data_start_row = 3
        current_row = data_start_row

        # 遍历所有评估点
        for item_id, item_data in analysis['assessments'].items():
            # 查找评估点（假设在第1列）
            if current_row < ws.max_row:
                # 更新评估结果（假设在第4列）
                result_cell = ws.cell(current_row, 4)
                result_cell.value = item_data['result']

                # 设置结果颜色
                if item_data['result'] == '符合':
                    result_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                elif item_data['result'] == '不符合':
                    result_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif item_data['result'] == '部分符合':
                    result_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                # 更新评估说明（假设在第5列）
                ws.cell(current_row, 5, item_data['explanation'])

                current_row += 1

        # 保存文件
        if not output_path:
            output_path = self.assessment_excel

        wb.save(output_path)
        logger.info(f"自评估表已更新: {output_path}")


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='增强版 AI Excel 自动填写工具')
    parser.add_argument('--result-dir', default='privacy_check_results', help='检查结果目录')
    parser.add_argument('--base-path', help='基础路径（包含 Excel 模板）')
    parser.add_argument('--enable-ai', action='store_true', help='启用 AI 分析')

    args = parser.parse_args()

    # 创建填充器
    filler = EnhancedAIExcelFiller(
        base_path=args.base_path,
        enable_ai=args.enable_ai
    )

    # 填写权限确认单
    filler.fill_permission_confirmation(args.result_dir)

    # 填写自评估表
    filler.fill_self_assessment(args.result_dir)

    print("\n✅ Excel 文件填写完成")


if __name__ == '__main__':
    main()
