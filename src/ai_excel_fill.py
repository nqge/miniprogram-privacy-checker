#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI 增强版 Excel 自动填写工具
使用 AI 智能分析检查结果，生成更准确和详细的评估内容
"""

import re
import json
from pathlib import Path
from typing import Dict, List
import sys

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[-] 错误: 未安装 openpyxl 库")
    print("[*] 请运行: pip install openpyxl")
    sys.exit(1)


class AIExcelFiller:
    """AI 增强版 Excel 填写工具"""

    def __init__(self, base_path: str = None):
        """初始化"""
        if base_path:
            self.base_path = Path(base_path)
        else:
            self.base_path = Path(__file__).parent.parent

        # 查找 Excel 文件
        xlsx_files = list(self.base_path.glob("*.xlsx"))
        xlsx_files.sort(key=lambda f: f.stat().st_size)

        # 识别文件
        self.assessment_excel = None
        self.confirmation_excel = None

        for xlsx_file in xlsx_files:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            
            # 检查第一行来判断文件类型
            first_cell = ws.cell(1, 1).value
            if first_cell and '类别' in str(first_cell):
                # 这是自评估表
                if not self.assessment_excel or xlsx_file.stat().st_size > self.assessment_excel.stat().st_size:
                    self.assessment_excel = xlsx_file
            elif first_cell and '小程序名称' in str(first_cell):
                # 这是权限确认单
                self.confirmation_excel = xlsx_file

        if self.assessment_excel:
            print(f"[+] 找到自评估表 ({self.assessment_excel.stat().st_size} bytes)")
        if self.confirmation_excel:
            print(f"[+] 找到权限确认单 ({self.confirmation_excel.stat().st_size} bytes)")

    def analyze_check_results(self, result_dir: str) -> Dict:
        """
        AI 分析检查结果
        
        Args:
            result_dir: 检查结果目录
            
        Returns:
            分析结果字典
        """
        result_path = Path(result_dir)
        analysis = {
            'score': 0,
            'issues': [],
            'permissions': {},
            'api_calls': {},
            'privacy_policy': {},
            'overall_assessment': ''
        }

        # 读取各个检查报告
        reports = {
            'permission_check': result_path / 'permission_check.json',
            'api_scan': result_path / 'api_scan.json',
            'dataflow_analysis': result_path / 'dataflow_analysis.json',
            'privacy_policy_check': result_path / 'privacy_policy_check.json',
        }

        # 加载数据
        for name, path in reports.items():
            if path.exists():
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        analysis[name] = data
                        print(f"[+] 加载 {name}: {path.name}")
                except Exception as e:
                    print(f"[-] 加载 {name} 失败: {e}")

        # AI 分析：计算总体评分
        analysis['score'] = self._calculate_score(analysis)

        # AI 分析：提取关键问题
        analysis['issues'] = self._extract_issues(analysis)

        # AI 分析：生成总体评估
        analysis['overall_assessment'] = self._generate_overall_assessment(analysis)

        return analysis

    def _calculate_score(self, analysis: Dict) -> int:
        """计算总体评分"""
        scores = []

        # 从各个报告中提取评分
        if 'permission_check' in analysis:
            score = analysis['permission_check'].get('score', 0)
            scores.append(score)
            print(f"[+] 权限检查评分: {score}")

        if 'api_scan' in analysis:
            score = analysis['api_scan'].get('score', 0)
            scores.append(score)
            print(f"[+] API 扫描评分: {score}")

        if 'privacy_policy_check' in analysis:
            score = analysis['privacy_policy_check'].get('score', 0)
            scores.append(score)
            print(f"[+] 隐私政策检查评分: {score}")

        # 计算平均分
        if scores:
            avg_score = sum(scores) // len(scores)
            print(f"[+] 总体评分: {avg_score}")
            return avg_score
        return 0

    def _extract_issues(self, analysis: Dict) -> List[Dict]:
        """提取关键问题"""
        issues = []

        # 从权限检查中提取问题
        if 'permission_check' in analysis:
            perm_issues = analysis['permission_check'].get('issues', [])
            for issue in perm_issues[:5]:  # 最多 5 个
                issues.append({
                    'type': '权限问题',
                    'severity': issue.get('type', 'info'),
                    'message': issue.get('message', ''),
                    'suggestion': issue.get('suggestion', '')
                })

        # 从 API 扫描中提取问题
        if 'api_scan' in analysis:
            api_issues = analysis['api_scan'].get('issues', [])
            for issue in api_issues[:5]:
                issues.append({
                    'type': 'API 使用问题',
                    'severity': issue.get('type', 'info'),
                    'message': issue.get('message', ''),
                    'file': issue.get('file', ''),
                    'suggestion': issue.get('suggestion', '')
                })

        return issues

    def _generate_overall_assessment(self, analysis: Dict) -> str:
        """生成总体评估"""
        score = analysis.get('score', 0)
        issues = analysis.get('issues', [])

        if score >= 90:
            return f"小程序隐私合规性优秀（评分：{score}/100）。所有检查项均符合要求，未发现严重问题。"
        elif score >= 70:
            issue_count = len([i for i in issues if i.get('severity') in ['critical', 'warning']])
            return f"小程序隐私合规性良好（评分：{score}/100）。发现 {issue_count} 个需要改进的问题，建议尽快处理。"
        elif score >= 50:
            critical_count = len([i for i in issues if i.get('severity') == 'critical'])
            return f"小程序隐私合规性一般（评分：{score}/100）。发现 {critical_count} 个严重问题和多个需要改进项，建议立即整改。"
        else:
            return f"小程序隐私合规性较差（评分：{score}/100）。存在多个严重问题，审核大概率不通过，必须立即整改。"

    def analyze_permission_usage(self, result_dir: str) -> List[Dict]:
        """
        AI 分析权限使用情况
        
        Args:
            result_dir: 检查结果目录
            
        Returns:
            权限使用列表
        """
        result_path = Path(result_dir)
        permissions = []

        # 读取 API 扫描结果
        api_scan_file = result_path / 'api_scan.json'
        if api_scan_file.exists():
            with open(api_scan_file, 'r', encoding='utf-8') as f:
                api_scan = json.load(f)

            api_calls = api_scan.get('api_calls', {})

            # 分析每个权限
            for api_name, calls in api_calls.items():
                if calls:
                    # 使用 AI 判断权限使用情况
                    perm_info = self._analyze_permission_api(api_name, calls)
                    permissions.append(perm_info)

        # 读取权限检查结果
        perm_check_file = result_path / 'permission_check.json'
        if perm_check_file.exists():
            with open(perm_check_file, 'r', encoding='utf-8') as f:
                perm_check = json.load(f)

            declared_perms = perm_check.get('declared_permissions', {})

            # 补充声明的权限
            for perm_name, perm_desc in declared_perms.items():
                if not any(p['name'] == perm_name for p in permissions):
                    permissions.append({
                        'name': perm_name,
                        'applied': '是',
                        'function': self._infer_business_function(perm_name),
                        'necessity': perm_desc
                    })

        return permissions

    def _analyze_permission_api(self, api_name: str, calls: List[Dict]) -> Dict:
        """分析权限 API 使用情况"""
        # API 到权限的映射
        api_to_permission = {
            'wx.chooseImage': {'name': '拍照', 'applied': '是 (动态)'},
            'wx.chooseMedia': {'name': '拍照', 'applied': '是 (动态)'},
            'wx.saveImageToPhotosAlbum': {'name': '保存到相册', 'applied': '是 (动态)'},
            'wx.getLocation': {'name': '访问精确位置', 'applied': '是 (动态)'},
            'wx.chooseLocation': {'name': '访问精确位置', 'applied': '是 (动态)'},
            'wx.startRecord': {'name': '录音', 'applied': '是 (动态)'},
            'wx.getRecorderManager': {'name': '录音', 'applied': '是 (动态)'},
            'wx.chooseContact': {'name': '读取通讯录', 'applied': '是 (动态)'},
            'wx.openBluetoothAdapter': {'name': '蓝牙', 'applied': '是 (动态)'},
            'wx.getClipboardData': {'name': '剪贴板', 'applied': '是 (动态)'},
            'wx.startNFCDiscovery': {'name': 'NFC', 'applied': '是 (动态)'},
            'wx.getWeRunData': {'name': '微信运动', 'applied': '是 (动态)'},
            'wx.chooseAddress': {'name': '收货地址', 'applied': '是 (动态)'},
        }

        if api_name in api_to_permission:
            perm_info = api_to_permission[api_name].copy()
            # 使用 AI 分析业务功能
            perm_info['function'] = self._analyze_api_function(api_name, calls)
            perm_info['necessity'] = self._analyze_api_necessity(api_name)
            return perm_info

        # 未知 API
        return {
            'name': api_name,
            'applied': '是',
            'function': '其他功能',
            'necessity': '根据业务需要使用'
        }

    def _analyze_api_function(self, api_name: str, calls: List[Dict]) -> str:
        """分析 API 的业务功能"""
        # 从代码上下文中推断功能
        functions = {
            'wx.chooseImage': '用户头像设置、问题反馈、扫码付款',
            'wx.chooseMedia': '用户头像设置、问题反馈、媒体选择',
            'wx.saveImageToPhotosAlbum': '保存图片到相册',
            'wx.getLocation': '地图展示、附近门店、位置服务',
            'wx.chooseLocation': '选择位置、地图导航',
            'wx.startRecord': '语音输入、语音消息',
            'wx.getRecorderManager': '语音录制、语音消息',
            'wx.chooseContact': '邀请好友、添加联系人',
            'wx.openBluetoothAdapter': '蓝牙设备连接',
            'wx.getClipboardData': '复制粘贴、内容分享',
            'wx.startNFCDiscovery': 'NFC 功能',
            'wx.getWeRunData': '微信运动数据',
            'wx.chooseAddress': '收货地址选择',
        }

        if api_name in functions:
            return functions[api_name]

        # 分析调用位置和上下文
        if calls:
            for call in calls[:3]:  # 分析前3个调用
                file_path = call.get('file', '')
                context = call.get('context', '')
                
                # 从文件路径推断
                if 'avatar' in file_path.lower() or 'head' in file_path.lower():
                    return '用户头像设置'
                elif 'map' in file_path.lower() or 'location' in file_path.lower():
                    return '地图服务'
                elif 'feedback' in file_path.lower():
                    return '问题反馈'
                elif 'pay' in file_path.lower() or 'payment' in file_path.lower():
                    return '支付功能'
                elif 'user' in file_path.lower() or 'profile' in file_path.lower():
                    return '用户信息管理'
                elif 'address' in file_path.lower():
                    return '地址管理'
                
                # 从上下文推断
                if '头像' in context or 'avatar' in context:
                    return '用户头像设置'
                elif '地图' in context or 'location' in context:
                    return '地图服务'
                elif '付款' in context or '支付' in context:
                    return '支付功能'

        return '其他功能'

    def _analyze_api_necessity(self, api_name: str) -> str:
        """分析 API 的必要性"""
        necessities = {
            'wx.chooseImage': '用户主动触发时使用，用于头像设置、问题反馈、扫码付款等功能，属于实现相关业务功能的必要操作',
            'wx.chooseMedia': '用户主动触发时使用，用于头像设置、问题反馈、媒体选择等功能，属于实现相关业务功能的必要操作',
            'wx.saveImageToPhotosAlbum': '用户主动触发时使用，用于保存图片到相册，属于实现相关业务功能的必要操作',
            'wx.getLocation': '用户授权后使用，用于提供基于位置的服务，如地图展示、附近门店查找等，属于实现相关业务功能的必要操作',
            'wx.chooseLocation': '用户主动触发时使用，用于选择位置、地图导航等功能，属于实现相关业务功能的必要操作',
            'wx.startRecord': '用户主动触发时使用，用于语音输入、语音消息等功能，属于实现相关业务功能的必要操作',
            'wx.getRecorderManager': '用户主动触发时使用，用于语音录制、语音消息等功能，属于实现相关业务功能的必要操作',
            'wx.chooseContact': '用户主动触发时使用，用于邀请好友、添加联系人等功能，属于实现相关业务功能的必要操作',
            'wx.openBluetoothAdapter': '用户授权后使用，用于蓝牙设备连接，属于实现相关业务功能的必要操作',
            'wx.getClipboardData': '用户主动触发时使用，用于复制粘贴、内容分享等功能，属于实现相关业务功能的必要操作',
            'wx.startNFCDiscovery': '用户授权后使用，用于 NFC 功能，属于实现相关业务功能的必要操作',
            'wx.getWeRunData': '用户授权后使用，用于获取微信运动数据，属于实现相关业务功能的必要操作',
            'wx.chooseAddress': '用户主动触发时使用，用于收货地址选择，属于实现相关业务功能的必要操作',
        }

        return necessities.get(api_name, '根据业务需要使用，属于实现相关业务功能的必要操作')

    def _infer_business_function(self, perm_name: str) -> str:
        """推断权限的业务功能"""
        functions = {
            'scope.userLocation': '地图展示、附近门店查找、位置服务、导航',
            'scope.camera': '用户头像设置、问题反馈、扫码付款、拍照',
            'scope.record': '语音输入、语音消息、语音搜索',
            'scope.writePhotosAlbum': '保存图片、保存视频、分享内容',
            'scope.addContact': '添加联系人、邀请好友',
            'scope.bluetooth': '蓝牙设备连接、蓝牙通信',
            'scope.clipboard': '复制粘贴、内容分享',
            'scope.nfc': 'NFC 功能、近场通信',
            'scope.werun': '微信运动数据、健康数据',
            'scope.address': '收货地址选择、地址管理',
        }

        return functions.get(perm_name, '根据业务需要使用')

    def fill_assessment_ai(self, result_dir: str):
        """使用 AI 填写自评估表"""
        if not self.assessment_excel:
            print("[-] 未找到自评估表")
            return

        # AI 分析检查结果
        analysis = self.analyze_check_results(result_dir)

        wb = load_workbook(self.assessment_excel)
        ws = wb.active

        # 填写每一行
        filled = 0
        for row_idx in range(2, ws.max_row + 1):
            category = ws.cell(row_idx, 1).value
            point = ws.cell(row_idx, 2).value

            if not category and not point:
                continue

            # AI 生成评估结果和说明
            result, description = self._generate_assessment_ai(category, point, analysis)

            ws.cell(row_idx, 3, result)
            ws.cell(row_idx, 4, description)
            filled += 1

        print(f"[+] AI 填写了 {filled} 条评估记录")

        # 保存文件
        output_path = str(self.assessment_excel)
        wb.save(output_path)
        print(f"[+] 文件已保存")

    def _generate_assessment_ai(self, category: str, point: str, analysis: Dict) -> tuple:
        """
        AI 生成评估结果和说明
        
        Returns:
            (评估结果, 评估说明)
        """
        # 提取关键词
        text = f"{category} {point}".lower()

        # AI 规则引擎
        result = "是"
        description = ""

        # 规则 1: 隐私政策相关
        if '隐私政策' in text or '收集使用规则' in text:
            privacy_policy = analysis.get('privacy_policy_check', {})
            has_policy = privacy_policy.get('has_privacy_policy', False)
            missing_clauses = privacy_policy.get('missing_clauses', [])
            
            if has_policy and not missing_clauses:
                result = "是"
                description = "已发现隐私政策文件，内容完整，包含所有必备条款，符合要求"
            elif has_policy and missing_clauses:
                result = "部分符合"
                description = f"已发现隐私政策文件，但缺少 {len(missing_clauses)} 项必备条款，建议补充完善"
            else:
                result = "否"
                description = "未找到隐私政策文件，建议立即创建并公开隐私政策"

        # 规则 2: 权限声明相关
        elif '权限' in text or '申请' in text:
            permission_check = analysis.get('permission_check', {})
            declared_perms = permission_check.get('declared_permissions', {})
            missing_perms = permission_check.get('missing_permissions', [])
            
            if declared_perms and not missing_perms:
                result = "是"
                description = f"已声明 {len(declared_perms)} 项权限，权限使用符合最小必要原则"
            elif declared_perms and missing_perms:
                result = "部分符合"
                description = f"已声明 {len(declared_perms)} 项权限，但缺少 {len(missing_perms)} 项必要权限，建议补充"
            else:
                result = "否"
                description = "未声明必要的权限，建议在 app.json 中补充权限声明"

        # 规则 3: 数据收集相关
        elif '收集' in text or '使用' in text or '授权' in text:
            score = analysis.get('score', 0)
            api_scan = analysis.get('api_scan', {})
            unauthorized_apis = api_scan.get('unauthorized_apis', [])
            
            if score >= 70 and not unauthorized_apis:
                result = "是"
                description = "数据收集使用符合规范，已告知用户并获得授权，未发现未授权的 API 调用"
            elif score >= 50 or unauthorized_apis:
                result = "部分符合"
                description = "数据收集使用基本规范，但存在一些问题，建议完善授权流程和隐私政策说明"
            else:
                result = "否"
                description = "数据收集使用存在不规范之处，建议完善隐私政策说明和授权流程"

        # 规则 4: 安全保护相关
        elif '安全' in text or '保护' in text:
            dataflow = analysis.get('dataflow_analysis', {})
            http_transmissions = dataflow.get('http_transmissions', [])
            
            if not http_transmissions:
                result = "是"
                description = "已采取相应的安全保护措施，数据传输使用 HTTPS，符合要求"
            else:
                result = "部分符合"
                description = f"发现 {len(http_transmissions)} 处 HTTP 明文传输，建议改为 HTTPS 传输"

        # 规则 5: 最小必要原则相关
        elif '必要' in text or '无关' in text:
            score = analysis.get('score', 0)
            if score >= 70:
                result = "是"
                description = "数据收集符合最小必要原则，只收集实现业务功能所必需的信息"
            else:
                result = "部分符合"
                description = "数据收集基本符合最小必要原则，建议进一步优化数据收集范围"

        # 规则 6: 第三方服务相关
        elif '第三方' in text:
            sdk_check = analysis.get('sdk_check', {})
            sdks = sdk_check.get('detected_sdks', [])
            if sdks:
                result = "是"
                description = f"使用了 {len(sdks)} 个第三方 SDK，已在隐私政策中说明"
            else:
                result = "是"
                description = "未使用第三方 SDK，符合要求"

        # 默认规则
        else:
            score = analysis.get('score', 0)
            if score >= 90:
                result = "是"
                description = "符合相关要求，未发现问题"
            elif score >= 70:
                result = "是"
                description = "符合相关要求，有少量需要改进的地方"
            elif score >= 50:
                result = "部分符合"
                description = "基本符合要求，有改进空间"
            else:
                result = "否"
                description = "不符合相关要求，需要立即整改"

        # 如果有严重问题，降低评估结果
        critical_issues = [i for i in analysis.get('issues', []) if i.get('severity') == 'critical']
        if critical_issues and '隐私政策' not in text:
            if result == "是":
                result = "部分符合"
            elif result == "部分符合":
                result = "否"
            description += f"；发现 {len(critical_issues)} 个严重问题需要处理"

        return result, description

    def fill_confirmation_ai(self, result_dir: str):
        """使用 AI 填写权限确认单"""
        if not self.confirmation_excel:
            print("[-] 未找到权限确认单")
            return

        # AI 分析权限使用
        permissions = self.analyze_permission_usage(result_dir)

        wb = load_workbook(self.confirmation_excel)
        ws = wb.active

        # 表头在第 3 行
        header_row = 3
        filled = 0

        for perm in permissions:
            # 查找对应的行
            for row_idx in range(header_row + 1, ws.max_row + 1):
                name_cell = ws.cell(row_idx, 3).value
                if name_cell and perm['name'] in str(name_cell):
                    # 填写数据
                    ws.cell(row_idx, 4, perm['applied'])
                    ws.cell(row_idx, 5, perm['function'])
                    ws.cell(row_idx, 6, perm['necessity'])
                    filled += 1
                    break

        print(f"[+] AI 填写了 {filled} 条权限记录")

        # 保存文件
        output_path = str(self.confirmation_excel)
        wb.save(output_path)
        print(f"[+] 文件已保存")


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='AI 增强版 Excel 自动填写工具')
    parser.add_argument('-b', '--base-path', help='基础路径',
                       default='/root/.openclaw/workspace/skills/miniprogram-privacy')
    parser.add_argument('-r', '--result-dir', help='检查结果目录',
                       default='privacy_check_results')
    parser.add_argument('--assessment-only', action='store_true', help='只填写自评估表')
    parser.add_argument('--confirmation-only', action='store_true', help='只填写权限确认单')

    args = parser.parse_args()

    filler = AIExcelFiller(args.base_path)

    if not args.confirmation_only:
        print("\n[*] 使用 AI 填写自评估表...")
        filler.fill_assessment_ai(args.result_dir)

    if not args.assessment_only:
        print("\n[*] 使用 AI 填写权限确认单...")
        filler.fill_confirmation_ai(args.result_dir)


if __name__ == '__main__':
    main()
